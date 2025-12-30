import json
import time
import sys
import io
import os
from io import BytesIO

# CRITICAL: Set UTF-8 encoding with error replacement BEFORE any output
os.environ["PYTHONUTF8"] = "1"
try:
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer,
            encoding="utf-8",
            errors="replace"
        )
    if hasattr(sys.stderr, 'buffer'):
        sys.stderr = io.TextIOWrapper(
            sys.stderr.buffer,
            encoding="utf-8",
            errors="replace"
        )
except (AttributeError, ValueError):
    pass

import requests
from config import (
    ERP_ATTENDANCE_URL,
    ERP_DEFAULT_SHIFT_ID,
    ERP_SHIFTS_BASE_URL,
    ERP_WORKLOG_URL,
    BASE_URL,
    ERP_NOTICES_URL,
)
from utils.network_checker import get_public_ip
from utils.logger import logger
from utils.excel_storage import read_local_storage
from utils.capture_types import CaptureArtifact


def safe_str(text):
    """Safely convert any text to ASCII-safe string, replacing problematic Unicode."""
    if text is None:
        return ""
    try:
        return str(text).encode("ascii", "replace").decode("ascii")
    except Exception:
        return "[UnicodeError]"


def safe_print(*args, **kwargs):
    """Safe print function that handles Unicode encoding errors."""
    try:
        safe_args = [safe_str(arg) for arg in args]
        print(*safe_args, **kwargs)
    except Exception:
        pass  # Silent fail to prevent EXE crash


class AttendanceAPI:
    def __init__(self, auth_api):
        self.auth_api = auth_api
        self.client_id = None
        self.selected_shift_id = None
        # If ERP account requires physical machine punch, app UI should disable punch actions
        self.machine_punch_required = False
        # Optional UI-provided IP override used in payloads/headers if present
        self.ip_override = None
        # Default geolocation (used if none provided by caller)
        self._default_latitude = "19.651836"
        self._default_longitude = "74.258816"

    def set_client(self, client_id: str):
        self.client_id = client_id
        self.selected_shift_id = None

    def set_shift(self, shift_id: str):
        self.selected_shift_id = shift_id

    def set_ip_override(self, ip: str | None):
        self.ip_override = (ip or "").strip() or None

    def fetch_user_profile(self, client_id: str | None = None, user_id: str | None = None):
        """
        Fetch the logged-in user's profile (first_name, last_name etc.).
        Uses: GET /users/client/{client_id}/user/{user_id}
        """
        # Prefer explicit values, then stored client_id / user_id, then values from JWT
        client_id = client_id or self.client_id or self.auth_api.get_client_id_from_token()
        user_id = user_id or getattr(self.auth_api, "user_id", None) or self.auth_api.get_user_id_from_token()

        if not client_id or not user_id:
            return False, {"message": "client_id and user_id are required"}, 400

        url = f"{BASE_URL}/users/client/{client_id}/user/{user_id}"
        try:
            logger.info(f"[UserProfile] Fetching profile: client_id={client_id}, user_id={user_id}, url={url}")
            resp = self.auth_api.authorized_request("GET", url)
        except Exception as exc:
            logger.error("Error calling user profile API: %s", exc, exc_info=True)
            return False, {"message": str(exc)}, 500

        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}

        # Print / log full payload so we can see exact structure
        try:
            logger.info(f"[UserProfile] HTTP {resp.status_code} payload: {data}")
            print(f"[UserProfile] HTTP {resp.status_code} payload: {data}")
        except Exception:
            pass
        return resp.ok, data, resp.status_code

    def fetch_shifts(self, client_id: str, page: int = 1, limit: int = 10):
        url = f"{ERP_SHIFTS_BASE_URL}/{client_id}"
        params = {"page": page, "limit": limit}
        resp = self.auth_api.authorized_request("GET", url, params=params)
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        return resp.ok, data, resp.status_code

    def fetch_user_shifts(self, client_id: str, user_id: str = None):
        """
        Fetch shifts assigned to a specific user for a client.
        Uses the endpoint: /shifts/client/{client_id}/users/{user_id}
        Response format: { "success": true, "message": "...", "shifts": [...] }
        """
        if not user_id:
            user_id = self.auth_api.user_id
        if not user_id:
            return False, {"message": "User ID is required"}, 400
        url = f"{ERP_SHIFTS_BASE_URL}/{client_id}/users/{user_id}"
        resp = self.auth_api.authorized_request("GET", url)
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        # Check for success field in response
        if resp.ok and isinstance(data, dict):
            # The API returns { "success": true, "message": "...", "shifts": [...] }
            # We return the data as-is, and the caller can extract shifts from data.get("shifts")
            return resp.ok, data, resp.status_code
        return resp.ok, data, resp.status_code

    # --- Shift helpers ---
    def _extract_shift_items(self, payload):
        items = []
        data = payload
        if isinstance(data, dict):
            for key in ("data", "items", "shifts", "rows", "result"):
                val = data.get(key)
                if isinstance(val, list) and val:
                    items = val
                    break
            if not items:
                items = [data]
        elif isinstance(data, list):
            items = data
        return [x for x in items if isinstance(x, dict)]

    def _shift_id_from(self, item: dict):
        return item.get("id") or item.get("shift_id") or item.get("_id")

    def auto_select_shift(self, client_id: str, use_user_shifts: bool = True) -> str | None:
        """
        Auto-select a shift for the user. 
        If use_user_shifts is True, tries fetch_user_shifts first (for assigned shifts).
        Falls back to fetch_shifts if user shifts fails or returns no shifts.
        """
        # Try user shifts first (assigned shifts) if user_id is available
        if use_user_shifts and self.auth_api.user_id:
            ok, payload, _ = self.fetch_user_shifts(client_id, self.auth_api.user_id)
            if ok:
                items = self._extract_shift_items(payload)
                if items:
                    # Prefer current shift if available
                    for it in items:
                        if it.get("is_current", False):
                            sid = self._shift_id_from(it)
                            if sid:
                                self.selected_shift_id = sid
                                return sid
                    # Otherwise use first shift from user shifts
                    sid = self._shift_id_from(items[0])
                    if sid:
                        self.selected_shift_id = sid
                        return sid
        
        # Fallback to regular shifts endpoint
        ok, payload, _ = self.fetch_shifts(client_id)
        if not ok:
            return None
        items = self._extract_shift_items(payload)
        if not items:
            return None
        def is_assigned(obj: dict) -> bool:
            for k in ("is_assigned", "assigned", "assigned_to_user", "assignedToUser"):
                v = obj.get(k)
                if isinstance(v, bool) and v:
                    return True
                if isinstance(v, (int, str)) and str(v).lower() in ("1", "true", "yes"):
                    return True
            return False
        # Prefer assigned
        for it in items:
            if is_assigned(it):
                sid = self._shift_id_from(it)
                if sid:
                    self.selected_shift_id = sid
                    return sid
        # Single or first
        if len(items) == 1:
            sid = self._shift_id_from(items[0])
            if sid:
                self.selected_shift_id = sid
                return sid
        for it in items:
            sid = self._shift_id_from(it)
            if sid:
                self.selected_shift_id = sid
                return sid
        return None

    def fetch_attendance(self, client_id: str, shift_id: str = None, date: str = None, month: int = None, year: int = None):
        params = {"client_id": client_id}
        if shift_id:
            params["shift_id"] = shift_id
        if date:
            params["date"] = date
        if month:
            params["month"] = month
        if year:
            params["year"] = year
        resp = self.auth_api.authorized_request("GET", ERP_ATTENDANCE_URL, params=params)
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        return resp.ok, data, resp.status_code

    def fetch_notices(self, client_id: str, page: int = 1, limit: int = 10):
        """
        Fetch announcement/notice list for the given client.
        Endpoint: /notices/client/{client_id}/notices
        """
        url = f"{ERP_NOTICES_URL}/{client_id}/notices"
        params = {"page": page, "limit": limit}
        # Use ERP default headers to mimic browser (origin/referer)
        headers = {}
        try:
            from config import ERP_DEFAULT_HEADERS
            headers.update(ERP_DEFAULT_HEADERS)
        except Exception:
            pass
        resp = self.auth_api.authorized_request("GET", url, params=params, headers=headers)
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        return resp.ok, data, resp.status_code

    def log_event(self, action, timestamp, metadata=None):
        payload = {
            "action": action,
            "timestamp": timestamp.isoformat(),
            "metadata": metadata or {}
        }
        headers = {"content-type": "application/json"}
        self.auth_api.authorized_request("POST", ERP_ATTENDANCE_URL, json=payload, headers=headers)

    def punch_in(self, client_id, date_in_iso_format, in_time_iso_utc, shift_id=None, latitude=None, longitude=None, status="W"):
        # Payload aligned to ERP API requirement
        if not shift_id:
            shift_id = self.selected_shift_id or ERP_DEFAULT_SHIFT_ID or None
        ip_addr = self.ip_override or ""
        if not ip_addr:
            try:
                ip_addr = get_public_ip()
            except Exception:
                ip_addr = ""
        payload = {
            "client_id": client_id,
            "action": "punch_in",
            "shift_id": shift_id,
            "date_in_iso_format": date_in_iso_format,
            "in_time": in_time_iso_utc,
            "status": status,
            "latitude": latitude if latitude is not None else self._default_latitude,
            "longitude": longitude if longitude is not None else self._default_longitude,
            "ip": ip_addr,
            "ip_address": ip_addr,
        }
        headers = {"content-type": "application/json", "x-client-ip": ip_addr or ""}
        resp = self.auth_api.authorized_request("POST", ERP_ATTENDANCE_URL, json=payload, headers=headers)
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        # Detect ERP policy error that mandates machine punch
        try:
            message_text = (data.get("message") if isinstance(data, dict) else str(data) if data else "")
            if isinstance(message_text, str) and "machine punch" in message_text.lower():
                self.machine_punch_required = True
        except Exception:
            pass
        return resp.ok, data, resp.status_code

    def punch_out(self, date_in_iso_format, out_time_iso_utc, status="W", shift_id=None, latitude=None, longitude=None):
        if not shift_id:
            shift_id = self.selected_shift_id or ERP_DEFAULT_SHIFT_ID or None
        ip_addr = self.ip_override or ""
        if not ip_addr:
            try:
                ip_addr = get_public_ip()
            except Exception:
                ip_addr = ""
        payload = {
            "client_id": self.client_id,
            "action": "punch_out",
            "shift_id": shift_id,
            "date_in_iso_format": date_in_iso_format,
            "out_time": out_time_iso_utc,
            "status": status,
            "latitude": latitude if latitude is not None else self._default_latitude,
            "longitude": longitude if longitude is not None else self._default_longitude,
            "ip": ip_addr,
            "ip_address": ip_addr,
        }
        headers = {"content-type": "application/json", "x-client-ip": ip_addr or ""}
        resp = self.auth_api.authorized_request("POST", ERP_ATTENDANCE_URL, json=payload, headers=headers)
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        return resp.ok, data, resp.status_code

    def start_break(self, date_in_iso_format, break_start_iso_utc):
        # ERP expects action "break" with key break_time
        ip_addr = self.ip_override or ""
        if not ip_addr:
            try:
                ip_addr = get_public_ip()
            except Exception:
                ip_addr = ""
        payload = {
            "client_id": self.client_id,
            "shift_id": self.selected_shift_id or ERP_DEFAULT_SHIFT_ID or None,
            "action": "break",
            "date_in_iso_format": date_in_iso_format,
            "break_time": break_start_iso_utc,
            "latitude": self._default_latitude,
            "longitude": self._default_longitude,
            "ip": ip_addr,
            "ip_address": ip_addr,
        }
        print(payload)
        headers = {"content-type": "application/json", "x-client-ip": ip_addr or ""}
        resp = self.auth_api.authorized_request("POST", ERP_ATTENDANCE_URL, json=payload, headers=headers)
        print(resp.json())
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        return resp.ok, data, resp.status_code

    def end_break(self, date_in_iso_format, break_end_iso_utc):
        # ERP expects a resumed action with resume_time
        ip_addr = self.ip_override or ""
        if not ip_addr:
            try:
                ip_addr = get_public_ip()
            except Exception:
                ip_addr = ""
        payload = {
            "client_id": self.client_id,
            "shift_id": self.selected_shift_id or ERP_DEFAULT_SHIFT_ID or None,
            "action": "resumed",
            "date_in_iso_format": date_in_iso_format,
            "resume_time": break_end_iso_utc,
            "latitude": self._default_latitude,
            "longitude": self._default_longitude,
            "ip": ip_addr,
            "ip_address": ip_addr,
        }
        print(payload)
        headers = {"content-type": "application/json", "x-client-ip": ip_addr or ""}
        resp = self.auth_api.authorized_request("POST", ERP_ATTENDANCE_URL, json=payload, headers=headers)
        try:
            data = resp.json()
        except Exception:
            data = {"message": resp.text}
        return resp.ok, data, resp.status_code

    def get_daily_summary(self):
        # Kept for backward compatibility; GUI prefers SessionManager for daily
        return {"total_active_time": "00:00:00", "total_break_time": "00:00:00", "total_idle_time": "00:00:00"}

    # Helpers for local aggregation
    def _load_history(self):
        try:
            data = read_local_storage()
            return data.get("history", [])
        except Exception:
            return []

    @staticmethod
    def _parse_hms(hms: str) -> int:
        try:
            parts = [int(p) for p in (hms or "00:00:00").split(":", 2)]
            while len(parts) < 3:
                parts.insert(0, 0)
            return parts[0] * 3600 + parts[1] * 60 + parts[2]
        except Exception:
            return 0

    @staticmethod
    def _format_hms(total_seconds: int) -> str:
        total_seconds = int(total_seconds)
        hrs = total_seconds // 3600
        mins = (total_seconds % 3600) // 60
        secs = total_seconds % 60
        return f"{hrs:02d}:{mins:02d}:{secs:02d}"

    def _aggregate(self, rows):
        active = sum(self._parse_hms(r.get("active")) for r in rows)
        brk = sum(self._parse_hms(r.get("break")) for r in rows)
        idle = sum(self._parse_hms(r.get("idle")) for r in rows)
        return {
            "total_active_time": self._format_hms(active),
            "total_break_time": self._format_hms(brk),
            "total_idle_time": self._format_hms(idle),
        }

    # The following aggregate from local history until server endpoints are available.
    def get_weekly_summary(self):
        try:
            from datetime import date, timedelta
            today = date.today()
            start = today - timedelta(days=6)  # last 7 days including today
            history = self._load_history()
            rows = [r for r in history if r.get("date") and start.isoformat() <= r["date"] <= today.isoformat()]
            return self._aggregate(rows)
        except Exception:
            return {"total_active_time": "00:00:00", "total_break_time": "00:00:00", "total_idle_time": "00:00:00"}

    def get_monthly_summary(self):
        try:
            from datetime import date
            today = date.today()
            month_prefix = today.strftime("%Y-%m")
            history = self._load_history()
            rows = [r for r in history if (r.get("date") or "").startswith(month_prefix)]
            return self._aggregate(rows)
        except Exception:
            return {"total_active_time": "00:00:00", "total_break_time": "00:00:00", "total_idle_time": "00:00:00"}

    def get_all_time_summary(self):
        try:
            history = self._load_history()
            return self._aggregate(history)
        except Exception:
            return {"total_active_time": "00:00:00", "total_break_time": "00:00:00", "total_idle_time": "00:00:00"}

    def get_history(self):
        # List of dicts with keys: date, active, break, idle, laptop_sleep
        try:
            data = read_local_storage()
            history = data.get("history", [])
            sleep_events = data.get("sleep_events", [])
            # Build map: date -> total sleep seconds
            sleep_by_date = {}
            for event in sleep_events:
                date = event.get("date")
                dur = int(event.get("duration_seconds", 0))
                if date:
                    sleep_by_date[date] = sleep_by_date.get(date, 0) + dur
            # Ensure the shape and stable ordering (latest first)
            normalized = []
            for item in history:
                dt = item.get("date")
                sleep_sec = sleep_by_date.get(dt, 0)
                laptop_sleep = self._format_hms(sleep_sec) if sleep_sec else "00:00:00"
                normalized.append({
                    "date": dt,
                    "active": item.get("active", "00:00:00"),
                    "break": item.get("break", "00:00:00"),
                    "idle": item.get("idle", "00:00:00"),
                    "laptop_sleep": laptop_sleep
                })
            # sort by date desc if parseable
            try:
                from datetime import datetime
                normalized.sort(key=lambda x: datetime.fromisoformat(x.get("date", "1970-01-01")), reverse=True)
            except Exception:
                pass
            return normalized
        except Exception:
            return []

    def upload_worklog_event(
        self,
        payload: dict,
        screenshot_artifact: CaptureArtifact | None = None,
        webcam_artifact: CaptureArtifact | None = None,
        tabs_snapshot: dict | None = None,
    ) -> bool:
        """
        Upload a combined worklog event (screenshot + webcam photo + browser tabs).
        """
        files = {}
        buffers = []
        body = dict(payload or {})
        if tabs_snapshot:
            body["tabs_snapshot"] = tabs_snapshot
        try:
            if screenshot_artifact and screenshot_artifact.has_payload():
                buf = BytesIO(screenshot_artifact.data)
                buffers.append(buf)
                files["screenshot"] = (
                    screenshot_artifact.filename,
                    buf,
                    screenshot_artifact.mimetype,
                )
            if webcam_artifact and webcam_artifact.has_payload():
                buf = BytesIO(webcam_artifact.data)
                buffers.append(buf)
                files["webcam"] = (
                    webcam_artifact.filename,
                    buf,
                    webcam_artifact.mimetype,
                )
        except Exception as exc:
            logger.error("Unable to prepare worklog attachments: %s", exc, exc_info=True)

        request_kwargs = {
            "data": {"payload": json.dumps(body)},
        }
        if files:
            request_kwargs["files"] = files

        try:
            resp = self.auth_api.authorized_request("POST", ERP_WORKLOG_URL, **request_kwargs)
            if not resp.ok:
                logger.warning("Worklog upload failed (%s): %s", resp.status_code, resp.text)
            return resp.ok
        except Exception as exc:
            logger.error("Worklog upload crashed: %s", exc, exc_info=True)
            return False
        finally:
            for buf in buffers:
                try:
                    buf.close()
                except Exception:
                    pass

    def upload_capture_asset(
        self,
        artifact: CaptureArtifact | None,
        capture_type: str = "screenshot",
    ) -> tuple[bool, str | None]:
        """
        Upload a single screenshot/webcam image to the ERP work-log endpoint.
        Returns (success, key) where key is the S3 key from the response.
        """
        if not artifact or not artifact.has_payload():
            return False, None
        if not self.client_id or not getattr(self.auth_api, "user_id", None):
            logger.warning("Capture upload skipped: missing client_id or user_id")
            return False, None

        url = f"{BASE_URL}/work-logs/client/{self.client_id}/user/{self.auth_api.user_id}/upload"
        buffers = []
        files = {}
        try:
            buf = BytesIO(artifact.data)
            buffers.append(buf)
            files["image"] = (artifact.filename, buf, artifact.mimetype)
            resp = self.auth_api.authorized_request("POST", url, files=files)
            if not resp.ok:
                logger.warning(
                    "Capture upload failed (%s): status=%s msg=%s",
                    capture_type,
                    resp.status_code,
                    resp.text,
                )
                return False, None
            data = {}
            try:
                data = resp.json()
            except Exception:
                pass
            key = self._extract_key_from_payload(data)
            logger.info(
                "Capture upload success (%s): filename=%s key=%s",
                capture_type,
                artifact.filename,
                key,
            )
            return True, key
        except Exception as exc:
            logger.error("Capture upload crashed (%s): %s", capture_type, exc, exc_info=True)
            return False, None
        finally:
            for buf in buffers:
                try:
                    buf.close()
                except Exception:
                    pass

    def _prepare_excel_for_upload(self, excel_path) -> tuple[bool, str | None]:
        """
        Prepare and validate Excel file before upload by ensuring all required fields are present.
        Returns (success, temp_file_path) where temp_file_path is a cleaned version for upload.
        """
        from openpyxl import load_workbook
        from pathlib import Path
        from tempfile import NamedTemporaryFile
        from config import EXCEL_ACTIVITY_SHEET
        from utils.excel_storage import ACTIVITY_HEADERS
        
        logger.info(f"[Excel Prep] Starting preparation of Excel file: {excel_path}")
        
        try:
            # Load the workbook
            logger.debug(f"[Excel Prep] Loading workbook...")
            wb = load_workbook(excel_path, data_only=False, keep_links=False)
            
            if EXCEL_ACTIVITY_SHEET not in wb.sheetnames:
                logger.warning(f"[Excel Prep] Sheet '{EXCEL_ACTIVITY_SHEET}' not found in workbook")
                wb.close()
                return False, None
            
            ws = wb[EXCEL_ACTIVITY_SHEET]
            logger.debug(f"[Excel Prep] Found sheet '{EXCEL_ACTIVITY_SHEET}' with {ws.max_row} rows")
            
            # Get headers
            if ws.max_row == 0:
                logger.warning(f"[Excel Prep] Workbook has no rows")
                wb.close()
                return False, None
            
            header_row = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
            # Create header map with case-insensitive and space-normalized matching
            header_map = {}
            for idx, h in enumerate(header_row):
                if h:
                    # Store original
                    header_map[h] = idx
                    # Also store normalized versions for matching
                    normalized = h.lower().replace(" ", "_").replace("-", "_")
                    if normalized not in header_map:
                        header_map[normalized] = idx
                    # Also store with spaces
                    space_version = h.replace("_", " ").lower()
                    if space_version not in header_map:
                        header_map[space_version] = idx
            
            logger.info(f"[Excel Prep] Headers found in Excel: {header_row}")
            print(f"   [INFO] Excel headers: {header_row}")
            
            # Get current user_id and client_id for filling missing values
            current_user_id = getattr(self.auth_api, "user_id", "") or "unknown"
            current_client_id = self.client_id or ""
            logger.info(f"[Excel Prep] Using user_id: {current_user_id}, client_id: {current_client_id}")
            
            # Required fields according to API - try multiple variations
            required_fields = ["user_id", "start_time", "activity_type", "end_time"]
            required_indices = {}
            
            # Field name variations the API might accept
            field_variations = {
                "user_id": ["user_id", "user id", "userid", "userId", "user-id"],
                "start_time": ["start_time", "start time", "starttime", "startTime", "start-time"],
                "activity_type": ["activity_type", "activity type", "activitytype", "activityType", "activity-type"],
                "end_time": ["end_time", "end time", "endtime", "endTime", "end-time"]
            }
            
            for field in required_fields:
                found = False
                # Try exact match first
                if field in header_map:
                    required_indices[field] = header_map[field]
                    logger.info(f"[Excel Prep] Found '{field}' at column index {header_map[field]} (exact match)")
                    found = True
                else:
                    # Try variations
                    for variation in field_variations.get(field, [field]):
                        normalized_var = variation.lower().replace(" ", "_").replace("-", "_")
                        if normalized_var in header_map:
                            required_indices[field] = header_map[normalized_var]
                            logger.info(f"[Excel Prep] Found '{field}' at column index {header_map[normalized_var]} (matched '{variation}')")
                            found = True
                            break
                        # Also try direct lookup
                        if variation in header_map:
                            required_indices[field] = header_map[variation]
                            logger.info(f"[Excel Prep] Found '{field}' at column index {header_map[variation]} (matched '{variation}')")
                            found = True
                            break
                
                # Fallback: try position-based if field is in ACTIVITY_HEADERS
                if not found and field in ACTIVITY_HEADERS:
                    try:
                        idx = ACTIVITY_HEADERS.index(field)
                        if idx < len(header_row):
                            required_indices[field] = idx
                            logger.info(f"[Excel Prep] Found '{field}' at position {idx} via ACTIVITY_HEADERS (fallback)")
                            found = True
                    except (ValueError, IndexError):
                        pass
                
                if not found:
                    logger.warning(f"[Excel Prep] Could not find column for '{field}' in Excel headers")
            
            # Check if we have all required columns
            missing_cols = [f for f in required_fields if f not in required_indices]
            if missing_cols:
                logger.warning(f"[Excel Prep] ERROR: Missing required columns: {missing_cols}")
                wb.close()
                return False, None
            
            logger.info(f"[Excel Prep] SUCCESS: All required columns found: {list(required_indices.keys())}")
            
            # Process rows and fix missing required fields
            rows_to_keep = []
            fixed_count = 0
            skipped_count = 0
            total_rows = ws.max_row - 1  # Exclude header row
            
            logger.info(f"[Excel Prep] Processing {total_rows} data rows...")
            
            for row_idx in range(2, ws.max_row + 1):
                row_cells = [cell for cell in ws[row_idx]]
                row_data = []
                for cell in row_cells:
                    value = cell.value
                    # Convert None to empty string, datetime to string
                    if value is None:
                        row_data.append("")
                    elif hasattr(value, 'isoformat'):  # datetime object
                        row_data.append(value.isoformat())
                    else:
                        row_data.append(str(value) if value else "")
                
                # Extract required fields with proper handling
                def get_field_value(field_name):
                    if field_name not in required_indices:
                        return None
                    idx = required_indices[field_name]
                    if idx < len(row_data):
                        val = row_data[idx]
                        # Return empty string if None or whitespace
                        if val is None:
                            return ""
                        val_str = str(val).strip()
                        return val_str if val_str else ""
                    return None
                
                user_id = get_field_value("user_id")
                start_time = get_field_value("start_time")
                activity_type = get_field_value("activity_type")
                end_time = get_field_value("end_time")
                
                # Check if row has any data at all
                if not any(str(v).strip() for v in row_data if v):
                    continue
                
                # Ensure client_id populated
                client_idx = header_map.get("client_id")
                if client_idx is not None and client_idx < len(row_data):
                    client_val = str(row_data[client_idx]).strip() if row_data[client_idx] else ""
                    if not client_val or client_val.lower() in ["unknown", "none", "null", ""]:
                        row_data[client_idx] = current_client_id
                else:
                    client_idx = None

                # Fix missing required fields
                needs_fix = False
                # Check if user_id is empty or "unknown" and replace with actual user_id
                if not user_id or (isinstance(user_id, str) and user_id.strip().lower() in ["unknown", "none", "null", ""]):
                    user_id = current_user_id
                    if "user_id" in required_indices:
                        row_data[required_indices["user_id"]] = user_id
                    needs_fix = True
                    logger.debug(f"[Excel Prep] Row {row_idx}: Fixed invalid user_id -> '{user_id}'")
                
                if not start_time:
                    # Skip rows without start_time - can't fix this
                    skipped_count += 1
                    logger.debug(f"[Excel Prep] Row {row_idx}: Skipped (missing start_time)")
                    continue
                
                if not activity_type:
                    # Use action or tool as activity_type if available
                    action_idx = header_map.get("action", header_map.get("action", -1))
                    tool_idx = header_map.get("tool", header_map.get("tool", -1))
                    if action_idx >= 0 and action_idx < len(row_data) and row_data[action_idx]:
                        activity_type = str(row_data[action_idx]).strip()
                    elif tool_idx >= 0 and tool_idx < len(row_data) and row_data[tool_idx]:
                        activity_type = str(row_data[tool_idx]).strip()
                    else:
                        activity_type = "active"  # Default
                    if "activity_type" in required_indices:
                        row_data[required_indices["activity_type"]] = activity_type
                    needs_fix = True
                    logger.debug(f"[Excel Prep] Row {row_idx}: Fixed missing activity_type -> '{activity_type}'")
                
                if not end_time:
                    # Set end_time to start_time if missing
                    end_time = start_time
                    if "end_time" in required_indices:
                        row_data[required_indices["end_time"]] = end_time
                    needs_fix = True
                    logger.debug(f"[Excel Prep] Row {row_idx}: Fixed missing end_time -> '{end_time}'")
                
                # Ensure all required fields have values
                if not user_id or not start_time or not activity_type or not end_time:
                    logger.warning(f"[Excel Prep] Row {row_idx}: Still missing required fields - user_id={bool(user_id)}, start_time={bool(start_time)}, activity_type={bool(activity_type)}, end_time={bool(end_time)}")
                    skipped_count += 1
                    continue
                
                # Ensure duration is set if missing
                duration_idx = header_map.get("duration", -1)
                if duration_idx >= 0:
                    duration = row_data[duration_idx] if duration_idx < len(row_data) else None
                    if not duration or (isinstance(duration, str) and not duration.strip()):
                        # Calculate duration from start_time and end_time
                        try:
                            from utils.excel_storage import _calculate_duration
                            duration = _calculate_duration(str(start_time), str(end_time))
                            if duration_idx < len(row_data):
                                row_data[duration_idx] = duration
                            needs_fix = True
                            logger.debug(f"[Excel Prep] Row {row_idx}: Calculated duration -> '{duration}'")
                        except Exception as e:
                            logger.debug(f"[Excel Prep] Row {row_idx}: Could not calculate duration: {e}")
                
                if needs_fix:
                    fixed_count += 1
                
                rows_to_keep.append(row_data)
            
            logger.info(f"[Excel Prep] Processing complete:")
            logger.info(f"[Excel Prep]   - Total rows processed: {total_rows}")
            logger.info(f"[Excel Prep]   - Valid rows to upload: {len(rows_to_keep)}")
            logger.info(f"[Excel Prep]   - Rows fixed: {fixed_count}")
            logger.info(f"[Excel Prep]   - Rows skipped: {skipped_count}")
            
            if not rows_to_keep:
                wb.close()
                logger.warning("[Excel Prep] ERROR: No valid rows to upload after validation")
                return False, None
            
            # Create a temporary cleaned workbook
            logger.debug("[Excel Prep] Creating temporary cleaned workbook...")
            temp_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_path = Path(temp_file.name)
            temp_file.close()
            logger.debug(f"[Excel Prep] Temporary file created: {temp_path}")
            
            # Create new workbook with cleaned data
            # IMPORTANT: Use API-expected column names (with underscores, no spaces)
            from openpyxl import Workbook
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = EXCEL_ACTIVITY_SHEET
            
            # Create API-compatible headers (ensure exact column names API expects)
            api_headers = []
            header_to_api_map = {}
            for orig_header in header_row:
                normalized = orig_header.lower().replace(" ", "_").replace("-", "_")
                # Map common variations to API expected names
                if normalized in ["user_id", "userid"]:
                    api_header = "user_id"
                elif normalized in ["start_time", "starttime"]:
                    api_header = "start_time"
                elif normalized in ["end_time", "endtime"]:
                    api_header = "end_time"
                elif normalized in ["activity_type", "activitytype"]:
                    api_header = "activity_type"
                else:
                    api_header = normalized
                api_headers.append(api_header)
                header_to_api_map[orig_header] = api_header

            uuid_column_indices = {
                header: idx
                for idx, header in enumerate(api_headers)
                if header.endswith("_id") and header not in {"client_id", "user_id"}
            }
            
            # Write API-compatible headers
            new_ws.append(api_headers)
            logger.info(f"[Excel Prep] API-compatible headers: {api_headers}")
            print(f"   [INFO] API headers: {api_headers}")
            
            # Create mapping from original header indices to API header indices
            orig_to_api_map = {}
            for orig_idx, orig_header in enumerate(header_row):
                if orig_idx < len(header_row):
                    api_header = header_to_api_map.get(orig_header, "")
                    if api_header and api_header in api_headers:
                        api_idx = api_headers.index(api_header)
                        orig_to_api_map[orig_idx] = api_idx
            
            # Write cleaned rows with proper column mapping
            for row_idx, row_data in enumerate(rows_to_keep, start=2):
                # Create a properly mapped row with all API columns
                mapped_row = [None] * len(api_headers)
                
                # Map data from original columns to API columns using the mapping
                for orig_idx, orig_value in enumerate(row_data):
                    if orig_idx in orig_to_api_map:
                        api_idx = orig_to_api_map[orig_idx]
                        if hasattr(orig_value, 'isoformat'):  # datetime
                            mapped_row[api_idx] = orig_value.isoformat()
                        elif isinstance(orig_value, str):
                            mapped_row[api_idx] = orig_value.strip() or None
                        elif orig_value is None:
                            mapped_row[api_idx] = None
                        else:
                            coerced = str(orig_value).strip()
                            mapped_row[api_idx] = coerced if coerced else None
                
                # Extract required field values from row_data using required_indices
                user_id_val = row_data[required_indices["user_id"]] if "user_id" in required_indices and required_indices["user_id"] < len(row_data) else ""
                start_time_val = row_data[required_indices["start_time"]] if "start_time" in required_indices and required_indices["start_time"] < len(row_data) else ""
                activity_type_val = row_data[required_indices["activity_type"]] if "activity_type" in required_indices and required_indices["activity_type"] < len(row_data) else ""
                end_time_val = row_data[required_indices["end_time"]] if "end_time" in required_indices and required_indices["end_time"] < len(row_data) else ""
                
                # Ensure required fields are set in mapped row
                if "client_id" in api_headers:
                    client_idx = api_headers.index("client_id")
                    client_val = mapped_row[client_idx]
                    if not client_val or (isinstance(client_val, str) and not client_val.strip()):
                        mapped_row[client_idx] = current_client_id or None

                if "user_id" in api_headers:
                    user_id_idx = api_headers.index("user_id")
                    user_id_str = str(user_id_val).strip() if user_id_val else ""
                    # Replace empty, "unknown", or invalid user_id with actual current_user_id
                    if not user_id_str or user_id_str.lower() in ["unknown", "none", "null", ""]:
                        user_id_str = current_user_id
                        logger.debug(f"[Excel Prep] Row {row_idx}: Replaced invalid user_id with current_user_id: {current_user_id}")
                    mapped_row[user_id_idx] = user_id_str
                
                start_time_str = ""
                if "start_time" in api_headers:
                    start_time_idx = api_headers.index("start_time")
                    start_time_str = str(start_time_val).strip() if start_time_val else ""
                    if hasattr(start_time_val, 'isoformat'):
                        start_time_str = start_time_val.isoformat()
                    mapped_row[start_time_idx] = start_time_str
                
                if "activity_type" in api_headers:
                    activity_type_idx = api_headers.index("activity_type")
                    activity_type_str = str(activity_type_val).strip() if activity_type_val else ""
                    if not activity_type_str:
                        # Try to get from action or tool
                        if "action" in api_headers:
                            action_idx = api_headers.index("action")
                            action_val = mapped_row[action_idx]
                            if isinstance(action_val, str) and action_val.strip():
                                activity_type_str = action_val.strip()
                        if not activity_type_str and "tool" in api_headers:
                            tool_idx = api_headers.index("tool")
                            tool_val = mapped_row[tool_idx]
                            if isinstance(tool_val, str) and tool_val.strip():
                                activity_type_str = tool_val.strip()
                        if not activity_type_str:
                            activity_type_str = "active"
                    mapped_row[activity_type_idx] = activity_type_str
                
                if "end_time" in api_headers:
                    end_time_idx = api_headers.index("end_time")
                    end_time_str = str(end_time_val).strip() if end_time_val else ""
                    if not end_time_str:
                        end_time_str = start_time_str  # Use start_time if missing
                    if hasattr(end_time_val, 'isoformat'):
                        end_time_str = end_time_val.isoformat()
                    mapped_row[end_time_idx] = end_time_str

                # Ensure optional UUID columns remain blank instead of empty strings
                for header, idx in uuid_column_indices.items():
                    value = mapped_row[idx]
                    if value is None:
                        continue
                    if isinstance(value, str):
                        cleaned = value.strip()
                        if not cleaned or cleaned.lower() in {"unknown", "none", "null"}:
                            mapped_row[idx] = None
                
                # Log first row for debugging with full details
                if row_idx == 2:
                    logger.info(f"[Excel Prep] Sample row data (row {row_idx}):")
                    logger.info(f"   Original row_data length: {len(row_data)}")
                    logger.info(f"   Mapped row length: {len(mapped_row)}")
                    logger.info(f"   user_id: '{mapped_row[api_headers.index('user_id')] if 'user_id' in api_headers else 'N/A'}' (type: {type(mapped_row[api_headers.index('user_id')] if 'user_id' in api_headers else None)})")
                    logger.info(f"   start_time: '{mapped_row[api_headers.index('start_time')] if 'start_time' in api_headers else 'N/A'}' (type: {type(mapped_row[api_headers.index('start_time')] if 'start_time' in api_headers else None)})")
                    logger.info(f"   activity_type: '{mapped_row[api_headers.index('activity_type')] if 'activity_type' in api_headers else 'N/A'}' (type: {type(mapped_row[api_headers.index('activity_type')] if 'activity_type' in api_headers else None)})")
                    logger.info(f"   end_time: '{mapped_row[api_headers.index('end_time')] if 'end_time' in api_headers else 'N/A'}' (type: {type(mapped_row[api_headers.index('end_time')] if 'end_time' in api_headers else None)})")
                    print(f"   [DEBUG] Sample row (row {row_idx}):")
                    print(f"      user_id: '{mapped_row[api_headers.index('user_id')] if 'user_id' in api_headers else 'N/A'}'")
                    print(f"      start_time: '{mapped_row[api_headers.index('start_time')] if 'start_time' in api_headers else 'N/A'}'")
                    print(f"      activity_type: '{mapped_row[api_headers.index('activity_type')] if 'activity_type' in api_headers else 'N/A'}'")
                    print(f"      end_time: '{mapped_row[api_headers.index('end_time')] if 'end_time' in api_headers else 'N/A'}'")
                    
                    # Also log the full mapped row to see all columns
                    logger.debug(f"[Excel Prep] Full mapped row: {mapped_row}")
                
                new_ws.append(mapped_row)
            
            new_wb.save(temp_path)
            new_wb.close()
            wb.close()
            
            # Verify the saved file by reading it back
            try:
                verify_wb = load_workbook(temp_path, data_only=True)
                verify_ws = verify_wb[EXCEL_ACTIVITY_SHEET]
                if verify_ws.max_row > 1:
                    verify_headers = [cell.value for cell in verify_ws[1]]
                    verify_row = [cell.value for cell in verify_ws[2]]  # First data row
                    logger.info(f"[Excel Prep] Verification - Headers: {verify_headers[:5]}...")
                    logger.info(f"[Excel Prep] Verification - First row sample: user_id='{verify_row[verify_headers.index('user_id')] if 'user_id' in verify_headers else 'N/A'}', start_time='{verify_row[verify_headers.index('start_time')] if 'start_time' in verify_headers else 'N/A'}'")
                    print(f"   [Verification] user_id='{verify_row[verify_headers.index('user_id')] if 'user_id' in verify_headers else 'N/A'}', start_time='{verify_row[verify_headers.index('start_time')] if 'start_time' in verify_headers else 'N/A'}'")
                verify_wb.close()
            except Exception as e:
                logger.warning(f"[Excel Prep] Could not verify saved file: {e}")
            
            file_size = temp_path.stat().st_size
            logger.info(f"[Excel Prep] SUCCESS: Prepared Excel file: {len(rows_to_keep)} valid rows, {file_size} bytes")
            return True, str(temp_path)
            
        except Exception as exc:
            logger.error(f"[Excel Prep] ERROR: Error preparing Excel file: {exc}", exc_info=True)
            try:
                wb.close()
            except Exception:
                pass
            return False, None

    def upload_activity_log_excel(self, client_id: str = None) -> tuple[bool, str]:
        """
        Upload the activity_log.xlsx file to the work-logs upload-logs endpoint.
        Validates and cleans the file before upload to ensure all required fields are present.
        Extracts client_id and user_id from JWT token if not provided.
        Returns (success, message).
        """
        logger.info("[Excel Upload] ===== Starting Excel upload process =====")
        safe_print("   [START] Starting Excel upload process...")
        
        # Try to extract client_id and user_id from JWT token
        token_client_id = self.auth_api.get_client_id_from_token()
        token_user_id = self.auth_api.get_user_id_from_token()
        
        # Use token values if available, otherwise fall back to stored values
        if not client_id:
            client_id = token_client_id or self.client_id
        
        if not client_id:
            logger.error("[Excel Upload] ERROR: Client ID is required (not found in token or stored)")
            safe_print("   ERROR: Client ID is required (not found in token or stored)")
            return False, "Client ID is required"
        
        # Log which source was used
        if token_client_id:
            logger.info(f"[Excel Upload] Client ID from token: {client_id}")
            safe_print(f"   [INFO] Client ID (from token): {client_id}")
        else:
            logger.info(f"[Excel Upload] Client ID from stored value: {client_id}")
            safe_print(f"   [INFO] Client ID (from stored): {client_id}")
        
        if token_user_id:
            logger.info(f"[Excel Upload] User ID from token: {token_user_id}")
            safe_print(f"   [INFO] User ID (from token): {token_user_id}")
        
        from pathlib import Path
        from config import EXCEL_ACTIVITY_FILE
        
        excel_path = Path(EXCEL_ACTIVITY_FILE)
        if not excel_path.exists():
            logger.warning(f"[Excel Upload] ERROR: Excel file not found: {excel_path}")
            return False, f"Excel file not found: {excel_path}"
        
        file_size = excel_path.stat().st_size
        logger.info(f"[Excel Upload] Found Excel file: {excel_path} ({file_size} bytes)")
        safe_print(f"   [FILE] Excel file: {excel_path.name} ({file_size:,} bytes)")
        
        # Prepare and validate Excel file
        logger.info("[Excel Upload] Step 1: Preparing and validating Excel file...")
        safe_print("   [STEP 1] Preparing and validating Excel file...")
        success, temp_file_path = self._prepare_excel_for_upload(excel_path)
        if not success or not temp_file_path:
            logger.error("[Excel Upload] ERROR: Failed to prepare Excel file for upload")
            safe_print("   ERROR: Failed to prepare Excel file for upload")
            return False, "Failed to prepare Excel file for upload (no valid rows or validation error)"
        
        logger.info(f"[Excel Upload] SUCCESS: Excel file prepared: {temp_file_path}")
        safe_print("   SUCCESS: Excel file prepared successfully")
        
        # Build the upload URL
        url = f"{BASE_URL}/work-logs/client/{client_id}/upload-logs"
        logger.info(f"[Excel Upload] Step 2: Uploading to URL: {url}")
        safe_print(f"   [STEP 2] Uploading to API...")
        
        temp_path = Path(temp_file_path)
        try:
            # Open the cleaned file and prepare for upload
            logger.debug(f"[Excel Upload] Opening temporary file for upload: {temp_path}")
            with open(temp_path, 'rb') as f:
                files = {
                    'file': (excel_path.name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                }
                logger.debug(f"[Excel Upload] File prepared: {excel_path.name}, size: {temp_path.stat().st_size} bytes")
                
                # Use authorized_request to handle authentication
                logger.info("[Excel Upload] Sending POST request to API...")
                resp = self.auth_api.authorized_request("POST", url, files=files)
                
                logger.info(f"[Excel Upload] Response received: Status {resp.status_code}")
                safe_print(f"   [RESPONSE] HTTP {resp.status_code}")
                
                if resp.ok:
                    logger.info("[Excel Upload] SUCCESS: Upload successful!")
                    safe_print("   SUCCESS: Upload successful!")
                    try:
                        data = resp.json()
                        message = data.get("message", "Upload successful") if isinstance(data, dict) else "Upload successful"
                        # CRITICAL: Sanitize message before any output or return
                        safe_message = safe_str(message)
                        logger.info(f"[Excel Upload] Response message: {safe_message}")
                        safe_print(f"   [INFO] Message: {safe_message}")
                        
                        # Log additional response details if available
                        if isinstance(data, dict):
                            if "success" in data:
                                logger.info(f"[Excel Upload] Success flag: {data.get('success')}")
                                safe_print(f"   [OK] Success flag: {data.get('success')}")
                            if "skippedRecords" in data:
                                skipped = data.get("skippedRecords", [])
                                if skipped:
                                    logger.warning(f"[Excel Upload] WARNING: {len(skipped)} records were skipped by API")
                                    safe_print(f"   WARNING: {len(skipped)} records were skipped by API")
                                    for record in skipped[:5]:  # Log first 5
                                        reason = safe_str(record.get('reason', ''))
                                        logger.warning(f"[Excel Upload]   - Row {record.get('row')}: {reason}")
                                        safe_print(f"      Row {record.get('row')}: {reason}")
                    except Exception as e:
                        logger.debug(f"[Excel Upload] Could not parse JSON response: {e}")
                        message = "Upload successful"
                    # Return sanitized message
                    return True, safe_str(message)
                else:
                    error_msg = resp.text
                    safe_error_msg = safe_str(error_msg[:500])  # Sanitize first 500 chars
                    logger.warning(f"[Excel Upload] ERROR: Upload failed with status {resp.status_code}")
                    logger.warning(f"[Excel Upload] Response text: {safe_error_msg}")
                    safe_print(f"   ERROR: Upload failed: HTTP {resp.status_code}")
                    
                    try:
                        data = resp.json()
                        error_msg = data.get("message", error_msg) if isinstance(data, dict) else error_msg
                        safe_error_msg = safe_str(error_msg)
                        logger.warning(f"[Excel Upload] Parsed error message: {safe_error_msg}")
                        
                        # Log skipped records if present
                        if isinstance(data, dict) and "skippedRecords" in data:
                            skipped = data.get("skippedRecords", [])
                            if skipped:
                                safe_print(f"   WARNING: {len(skipped)} records were skipped by API")
                                for record in skipped[:5]:
                                    reason = safe_str(record.get('reason', ''))
                                    safe_print(f"      Row {record.get('row')}: {reason}")
                        
                        # Return sanitized error message
                        return False, safe_str(error_msg)
                    except Exception:
                        pass
                    # Return sanitized error message
                    return False, safe_str(f"Upload failed: {error_msg}")
        except FileNotFoundError:
            logger.error(f"[Excel Upload] ERROR: Temporary file not found: {temp_path}")
            return False, safe_str(f"File not found: {temp_path}")
        except Exception as exc:
            logger.error(f"[Excel Upload] ERROR: Exception during upload: {exc}", exc_info=True)
            return False, safe_str(f"Upload error: {str(exc)}")
        finally:
            self._safe_delete_temp_file(temp_path)
        
        logger.info("[Excel Upload] ===== Upload process completed =====")

    def _safe_delete_temp_file(self, temp_path):
        """Retry temp file cleanup to avoid Windows locking issues."""
        if not temp_path:
            return
        attempts = 3
        for attempt in range(attempts):
            try:
                if temp_path.exists():
                    temp_path.unlink()
                    logger.debug(f"[Excel Upload] Cleaned up temporary file: {temp_path}")
                return
            except PermissionError as exc:
                if attempt < attempts - 1:
                    time.sleep(0.25)
                    continue
                logger.warning(f"[Excel Upload] Could not delete temporary file (locked): {exc}")
            except Exception as exc:
                logger.warning(f"[Excel Upload] Could not delete temporary file: {exc}")
                return

    @staticmethod
    def _extract_key_from_payload(payload) -> str | None:
        stack = [payload]
        while stack:
            current = stack.pop()
            if isinstance(current, dict):
                maybe_key = current.get("key")
                if isinstance(maybe_key, str) and maybe_key:
                    return maybe_key
                for value in current.values():
                    if isinstance(value, (dict, list)):
                        stack.append(value)
            elif isinstance(current, list):
                for item in current:
                    if isinstance(item, (dict, list)):
                        stack.append(item)
        return None
