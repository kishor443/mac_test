"""
Clean up temporary Excel files and fix the main file.
"""

from pathlib import Path
from utils.excel_storage import EXCEL_PATH
from openpyxl import load_workbook
import shutil

def cleanup_temp_files():
    """Clean up temp files and ensure main file is valid."""
    print("Cleaning up temporary files...")
    
    data_dir = EXCEL_PATH.parent
    temp_files = list(data_dir.glob("activity_log*.tmp.xlsx"))
    backup_files = list(data_dir.glob("activity_log*.backup.xlsx"))
    corrupted_files = list(data_dir.glob("activity_log*.corrupted*.xlsx"))
    
    all_temp = temp_files + backup_files + corrupted_files
    
    if all_temp:
        print(f"Found {len(all_temp)} temporary/backup files:")
        for f in all_temp:
            print(f"  - {f.name}")
        
        # Try to use temp file if it's newer and valid
        temp_file = data_dir / "activity_log.tmp.xlsx"
        if temp_file.exists():
            try:
                # Test if temp file is valid
                wb = load_workbook(temp_file, data_only=False, keep_links=False)
                wb.close()
                print("\n[INFO] Temp file is valid.")
                
                # Check if main file is locked
                try:
                    # Try to open main file in write mode to check if it's locked
                    test_wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
                    test_wb.close()
                    
                    # Main file is not locked, can replace
                    print("  Main file is not locked. Replacing with temp file...")
                    
                    # Backup current main file
                    if EXCEL_PATH.exists():
                        backup = EXCEL_PATH.with_suffix('.old.xlsx')
                        try:
                            shutil.copy2(EXCEL_PATH, backup)
                            print(f"  Backed up main file to: {backup.name}")
                        except Exception:
                            pass
                    
                    # Replace main file with temp
                    try:
                        shutil.copy2(temp_file, EXCEL_PATH)
                        print("  [OK] Main file replaced with temp file")
                        
                        # Delete temp file
                        temp_file.unlink()
                        print("  [OK] Temp file deleted")
                    except PermissionError:
                        print("  [WARNING] Cannot replace main file - it's locked by Excel.")
                        print("  Please close Excel and run this script again.")
                except PermissionError:
                    print("  [WARNING] Main file is locked (Excel is open).")
                    print("  Please close Excel and run this script again to replace the file.")
                    print(f"  Temp file saved at: {temp_file}")
            except Exception as e:
                print(f"  [ERROR] Temp file is invalid: {e}")
                try:
                    temp_file.unlink()
                    print("  [OK] Invalid temp file deleted")
                except Exception:
                    print("  [WARNING] Could not delete invalid temp file")
    
    # Clean up other temp/backup files
    for f in all_temp:
        if f.exists() and f != temp_file:
            try:
                f.unlink()
                print(f"  [OK] Deleted: {f.name}")
            except Exception as e:
                print(f"  [WARNING] Could not delete {f.name}: {e}")
    
    # Verify main file
    print("\nVerifying main file...")
    try:
        wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
        wb.close()
        print("[OK] Main Excel file is valid and ready to use!")
        print("\nPlease close Excel if it's open, then try opening the file again.")
    except Exception as e:
        print(f"[ERROR] Main file is still invalid: {e}")
        print("Run: python repair_excel.py to recreate the file")

if __name__ == "__main__":
    cleanup_temp_files()

