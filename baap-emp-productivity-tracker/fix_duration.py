"""
Fix duration column in existing Excel file to show proper time format.
"""

import sys
import os
import subprocess
import time
from datetime import datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.excel_storage import (
    EXCEL_PATH,
    ACTIVITY_HEADERS,
    EXCEL_ACTIVITY_SHEET,
    _calculate_duration,
    _set_column_widths,
)


def _parse_time(value):
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None

def fix_duration_column():
    """Fix duration column to show proper time format."""
    print("Fixing duration column...")
    print(f"File: {EXCEL_PATH}")
    
    # Close Excel first
    print("\n1. Closing Excel...")
    try:
        subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"], 
                      capture_output=True, timeout=15)
        time.sleep(2)
        print("   [OK] Excel closed")
    except Exception:
        print("   [INFO] Excel was not running")
    
    if not EXCEL_PATH.exists():
        print("[ERROR] File does not exist!")
        return False
    
    try:
        wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
        
        if EXCEL_ACTIVITY_SHEET not in wb.sheetnames:
            print(f"[ERROR] Sheet '{EXCEL_ACTIVITY_SHEET}' not found!")
            wb.close()
            return False
        
        ws = wb[EXCEL_ACTIVITY_SHEET]
        duration_col_idx = ACTIVITY_HEADERS.index("duration") + 1
        start_time_col_idx = ACTIVITY_HEADERS.index("start_time") + 1
        end_time_col_idx = ACTIVITY_HEADERS.index("end_time") + 1
        
        print(f"\n2. Processing {ws.max_row - 1} rows...")
        fixed_count = 0
        
        # Process each row
        for row_num in range(2, ws.max_row + 1):
            start_time = ws.cell(row=row_num, column=start_time_col_idx).value
            end_cell = ws.cell(row=row_num, column=end_time_col_idx)
            end_time = end_cell.value
            duration_cell = ws.cell(row=row_num, column=duration_col_idx)
            current_duration = duration_cell.value
            
            if not start_time:
                continue
            
            start_dt = _parse_time(start_time)
            end_dt = _parse_time(end_time)
            placeholder_end = (not end_time) or (str(end_time) == str(start_time))
            invalid_end = False
            if start_dt and end_dt and end_dt < start_dt:
                invalid_end = True
            
            if (placeholder_end or invalid_end) and start_dt:
                candidate = None
                if row_num < ws.max_row:
                    next_start = ws.cell(row=row_num + 1, column=start_time_col_idx).value
                    candidate = _parse_time(next_start)
                if candidate and candidate >= start_dt:
                    end_dt = candidate
                    end_cell.value = candidate
                else:
                    end_dt = start_dt
                    end_cell.value = start_time
            elif not end_dt and start_dt:
                end_dt = start_dt
                end_cell.value = start_time
            
            needs_fix = False
            if not current_duration or current_duration == "":
                needs_fix = True
            elif isinstance(current_duration, (int, float)):
                if current_duration < 24:
                    needs_fix = True
            elif not isinstance(current_duration, str) or ":" not in str(current_duration):
                needs_fix = True
            
            if needs_fix and start_dt and end_dt:
                try:
                    duration = _calculate_duration(str(start_dt), str(end_dt))
                    if duration:
                        duration_cell.value = duration
                        duration_cell.number_format = '@'  # Text format
                        fixed_count += 1
                except Exception:
                    pass
            elif current_duration:
                duration_cell.number_format = '@'
        
        # Set column widths
        _set_column_widths(ws, ACTIVITY_HEADERS)
        
        print(f"   [OK] Fixed {fixed_count} duration values")
        
        # Save
        print("\n3. Saving file...")
        wb.save(EXCEL_PATH)
        wb.close()
        
        print("\n" + "=" * 60)
        print("[SUCCESS] Duration column fixed!")
        print("=" * 60)
        print(f"\nFixed {fixed_count} rows")
        print("Duration format: HH:MM:SS (e.g., 00:05:30 for 5 minutes 30 seconds)")
        print(f"\nFile saved: {EXCEL_PATH.absolute()}")
        print("\nPlease close and reopen Excel to see updated durations.")
        
        return True
        
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_duration_column()

