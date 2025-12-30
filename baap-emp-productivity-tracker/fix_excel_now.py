"""
Quick fix for Excel file - closes Excel, cleans temp files, and repairs the file.
"""

import os
import sys
import time
import subprocess
from pathlib import Path

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from utils.excel_storage import EXCEL_PATH, ACTIVITY_HEADERS, EXCEL_ACTIVITY_SHEET, EXCEL_LOCAL_STORAGE_SHEET, LOCAL_STORAGE_HEADERS

def kill_excel_processes():
    """Kill all Excel processes on Windows."""
    try:
        subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"], 
                      capture_output=True, timeout=5)
        time.sleep(1)
        print("[OK] Closed Excel processes")
    except Exception:
        pass

def fix_excel_file():
    """Fix Excel file immediately."""
    print("=" * 60)
    print("FIXING EXCEL FILE - PLEASE WAIT...")
    print("=" * 60)
    
    # Step 1: Kill Excel processes
    print("\n1. Closing Excel...")
    kill_excel_processes()
    time.sleep(2)
    
    # Step 2: Clean temp files
    print("\n2. Cleaning temporary files...")
    data_dir = EXCEL_PATH.parent
    temp_patterns = ["*.tmp.xlsx", "*.backup.xlsx", "*.corrupted*.xlsx", "*.old.xlsx"]
    cleaned = 0
    for pattern in temp_patterns:
        for f in data_dir.glob(pattern):
            try:
                f.unlink()
                print(f"   Deleted: {f.name}")
                cleaned += 1
            except Exception:
                pass
    if cleaned == 0:
        print("   No temp files found")
    
    # Step 3: Check and repair main file
    print("\n3. Checking main file...")
    
    if not EXCEL_PATH.exists():
        print("   File doesn't exist. Creating new file...")
        wb = Workbook()
        ws = wb.active
        ws.title = EXCEL_ACTIVITY_SHEET
        ws.append(ACTIVITY_HEADERS)
        local_ws = wb.create_sheet(EXCEL_LOCAL_STORAGE_SHEET)
        local_ws.append(LOCAL_STORAGE_HEADERS)
        wb.save(EXCEL_PATH)
        print("[OK] New file created!")
        return True
    
    # Try to open existing file
    try:
        print("   Testing file...")
        wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
        
        # Verify structure
        if EXCEL_ACTIVITY_SHEET not in wb.sheetnames:
            print("   Missing ActivityLog sheet. Repairing...")
            wb.close()
            raise Exception("Missing sheet")
        
        wb.close()
        print("[OK] File is valid!")
        return True
        
    except (BadZipFile, InvalidFileException, KeyError, Exception) as e:
        print(f"   File is corrupted: {e}")
        print("   Creating backup and new file...")
        
        # Backup corrupted file
        try:
            backup = EXCEL_PATH.with_suffix('.corrupted_backup.xlsx')
            if EXCEL_PATH.exists():
                EXCEL_PATH.rename(backup)
                print(f"   Backed up to: {backup.name}")
        except Exception:
            try:
                EXCEL_PATH.unlink()
            except Exception:
                pass
        
        # Create new file
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = EXCEL_ACTIVITY_SHEET
            ws.append(ACTIVITY_HEADERS)
            local_ws = wb.create_sheet(EXCEL_LOCAL_STORAGE_SHEET)
            local_ws.append(LOCAL_STORAGE_HEADERS)
            wb.save(EXCEL_PATH)
            print("[OK] New file created!")
            return True
        except Exception as create_error:
            print(f"[ERROR] Could not create file: {create_error}")
            return False
    
    # Step 4: Final verification
    print("\n4. Final verification...")
    try:
        wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
        wb.close()
        print("[OK] File is ready to use!")
        print(f"\nFile location: {EXCEL_PATH.absolute()}")
        print("\nYou can now open the file in Excel.")
        return True
    except Exception as e:
        print(f"[ERROR] File still has issues: {e}")
        return False

if __name__ == "__main__":
    try:
        success = fix_excel_file()
        if success:
            print("\n" + "=" * 60)
            print("SUCCESS! Excel file is fixed and ready to use.")
            print("=" * 60)
        else:
            print("\n" + "=" * 60)
            print("ERROR: Could not fix file. Please check permissions.")
            print("=" * 60)
    except Exception as e:
        print(f"\n[FATAL ERROR] {e}")
        import traceback
        traceback.print_exc()

