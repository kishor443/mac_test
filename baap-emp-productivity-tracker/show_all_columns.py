"""
Fix Excel file to show all columns with proper widths.
"""

import sys
import os
import subprocess
import time
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils.excel_storage import (
    EXCEL_PATH, ACTIVITY_HEADERS, EXCEL_ACTIVITY_SHEET,
    _set_column_widths, _format_header_row
)

def fix_all_columns():
    """Ensure all columns are visible with proper widths."""
    print("Fixing Excel file to show all columns...")
    print(f"File: {EXCEL_PATH}")
    
    # Close Excel first
    print("\n1. Closing Excel...")
    try:
        subprocess.run(["taskkill", "/F", "/IM", "EXCEL.EXE"], 
                      capture_output=True, timeout=5)
        time.sleep(2)
        print("   [OK] Excel closed")
    except Exception:
        print("   [INFO] Excel was not running")
    
    if not EXCEL_PATH.exists():
        print("[ERROR] File does not exist!")
        return False
    
    try:
        # Load workbook
        wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
        
        if EXCEL_ACTIVITY_SHEET not in wb.sheetnames:
            print(f"[ERROR] Sheet '{EXCEL_ACTIVITY_SHEET}' not found!")
            wb.close()
            return False
        
        ws = wb[EXCEL_ACTIVITY_SHEET]
        
        # Get current headers
        current_headers = []
        if ws.max_row > 0:
            current_headers = [cell.value for cell in ws[1]]
        
        print(f"\nCurrent columns: {len(current_headers)}")
        print(f"Expected columns: {len(ACTIVITY_HEADERS)}")
        
        # Check if headers need updating
        if current_headers != ACTIVITY_HEADERS:
            print("\nUpdating headers...")
            
            # Save existing data
            existing_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    existing_data.append(list(row))
            
            # Clear sheet
            ws.delete_rows(1, ws.max_row)
            
            # Add new headers
            ws.append(ACTIVITY_HEADERS)
            _format_header_row(ws, ACTIVITY_HEADERS)
            _set_column_widths(ws, ACTIVITY_HEADERS)
            
            # Restore data with proper mapping
            print(f"Restoring {len(existing_data)} rows of data...")
            for row_data in existing_data:
                new_row = [""] * len(ACTIVITY_HEADERS)
                for idx, val in enumerate(row_data):
                    if idx < len(current_headers) and idx < len(ACTIVITY_HEADERS):
                        old_header = current_headers[idx]
                        if old_header in ACTIVITY_HEADERS:
                            new_idx = ACTIVITY_HEADERS.index(old_header)
                            new_row[new_idx] = val
                ws.append(new_row)
        else:
            print("\nHeaders are correct. Updating formatting...")
            _format_header_row(ws, ACTIVITY_HEADERS)
            _set_column_widths(ws, ACTIVITY_HEADERS)
        
        # Ensure no columns are hidden
        for col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].hidden = False
        
        # Set all column widths explicitly
        print("\nSetting column widths...")
        _set_column_widths(ws, ACTIVITY_HEADERS)
        
        # Freeze first row
        ws.freeze_panes = 'A2'
        
        # Save
        print("\nSaving file...")
        wb.save(EXCEL_PATH)
        wb.close()
        
        print("\n" + "=" * 60)
        print("[SUCCESS] All columns are now visible!")
        print("=" * 60)
        print(f"\nTotal columns: {len(ACTIVITY_HEADERS)}")
        print("Columns:")
        for i, header in enumerate(ACTIVITY_HEADERS, 1):
            print(f"  {i}. {header}")
        print(f"\nFile saved: {EXCEL_PATH.absolute()}")
        print("\nPlease close and reopen Excel to see all columns.")
        
        return True
        
    except Exception as e:
        print(f"\n[ERROR] {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    fix_all_columns()

