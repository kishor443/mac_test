"""
Test script to verify Excel file can be opened and is properly formatted.
"""

import os
import sys
from pathlib import Path

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from utils.excel_storage import EXCEL_PATH, ACTIVITY_HEADERS

def test_excel_file():
    """Test if Excel file can be opened and has correct structure."""
    print(f"Testing Excel file: {EXCEL_PATH}")
    print(f"File exists: {EXCEL_PATH.exists()}")
    
    if not EXCEL_PATH.exists():
        print("[ERROR] File does not exist!")
        return False
    
    print(f"File size: {EXCEL_PATH.stat().st_size} bytes")
    
    try:
        # Try to load workbook
        print("\nAttempting to load workbook...")
        wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
        print("[OK] Workbook loaded successfully!")
        
        # Check sheets
        print(f"\nSheets found: {wb.sheetnames}")
        
        # Check ActivityLog sheet
        if "ActivityLog" in wb.sheetnames:
            ws = wb["ActivityLog"]
            print(f"[OK] ActivityLog sheet found")
            print(f"  Rows: {ws.max_row}")
            print(f"  Columns: {ws.max_column}")
            
            # Check headers
            if ws.max_row > 0:
                headers = [cell.value for cell in ws[1]]
                print(f"  Headers: {headers[:5]}...")  # First 5 headers
                
                # Verify headers match expected
                if headers[:len(ACTIVITY_HEADERS)] == ACTIVITY_HEADERS:
                    print("[OK] Headers match expected format")
                else:
                    print("[WARNING] Headers don't match expected format")
        
        # Check for images
        if hasattr(wb, '_images') or hasattr(wb, 'images'):
            print("\n[INFO] Workbook contains embedded images")
        
        wb.close()
        print("\n[OK] Excel file is valid and can be opened!")
        print("\nIf Excel application still can't open it, try:")
        print("  1. Close Excel if it's open")
        print("  2. Run: python repair_excel.py")
        print("  3. Try opening the file again")
        return True
        
    except (BadZipFile, InvalidFileException) as e:
        print(f"\n[ERROR] File is corrupted: {e}")
        print("Run: python repair_excel.py to fix it")
        return False
    except PermissionError as e:
        print(f"\n[ERROR] Permission denied: {e}")
        print("Close Excel if it's open and try again")
        return False
    except Exception as e:
        print(f"\n[ERROR] Unexpected error: {e}")
        print("Run: python repair_excel.py to fix it")
        return False

if __name__ == "__main__":
    success = test_excel_file()
    sys.exit(0 if success else 1)

