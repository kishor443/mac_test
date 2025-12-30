from __future__ import annotations

import json
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
from io import BytesIO
from PIL import Image as PILImage

from config import (
    DATA_DIR,
    EXCEL_ACTIVITY_FILE,
    EXCEL_ACTIVITY_SHEET,
    EXCEL_LOCAL_STORAGE_SHEET,
)
from utils.capture_types import CaptureArtifact

EXCEL_PATH = Path(EXCEL_ACTIVITY_FILE)

ACTIVITY_HEADERS = [
    "client_id",
    "user_id",
    "tool",
    "action",
    "duration",
    "start_time",
    "end_time",
    "task_id",
    "project_id",
    "activity_type",
    "title",
    "metadata_json",
    "screenshots",
    "webcam_photo",
    "webcam_name",
    "mouse_clicks",
    "keys_count",
]

LOCAL_STORAGE_HEADERS = ["key", "json", "updated_at"]

_lock = threading.Lock()
_last_save_time = 0
_save_interval = 60.0  # Increased to 60 seconds to reduce disk I/O
_pending_rows = []  # Queue for batching writes
_last_activity_log_time = 0
_activity_log_interval = 10.0  # Minimum seconds between activity logs
_default_client_id = ""

# Background save executor to prevent UI blocking
from concurrent.futures import ThreadPoolExecutor
_save_executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="ExcelSaver")
_save_queue = []  # Queue of workbooks to save
_save_queue_lock = threading.Lock()


def _background_save_worker(wb, path):
    """Worker function that runs in background thread to save workbook."""
    try:
        wb.save(path)
        return True
    except Exception as e:
        try:
            from utils.logger import logger
            logger.error(f"Background save failed: {e}")
        except Exception:
            pass
        return False
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _backfill_client_and_user(client_id: str | None = None, user_id: str | None = None) -> None:
    """Update existing workbook rows to ensure client_id/user_id are populated."""
    if not client_id and not user_id:
        return
    wb = None
    try:
        with _lock:
            wb = _ensure_workbook()
            ws = _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
            client_idx = ACTIVITY_HEADERS.index("client_id") if client_id else None
            user_idx = ACTIVITY_HEADERS.index("user_id") if user_id else None
            updated = False
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row or 1):
                if client_idx is not None:
                    cell = row[client_idx]
                    value = str(cell.value).strip() if cell.value else ""
                    if not value or value.lower() in ("unknown", "none", "null"):
                        cell.value = client_id
                        updated = True
                if user_idx is not None:
                    cell = row[user_idx]
                    value = str(cell.value).strip() if cell.value else ""
                    if not value or value.lower() in ("unknown", "none", "null"):
                        cell.value = user_id
                        updated = True
            if updated:
                _save_workbook_with_retry(wb)
            else:
                # No changes, close immediately
                wb.close()
    except Exception:
        try:
            from utils.logger import logger
            logger.warning("Failed to backfill client/user IDs in Excel", exc_info=True)
        except Exception:
            pass
    finally:
        # Ensure workbook is closed even if error occurs
        if wb is not None:
            try:
                # Only close if not already submitted to background saver
                # Background saver will close it
                pass
            except Exception:
                pass


def set_default_client_id(client_id: str | None) -> None:
    """Allow caller (e.g., after client selection) to set fallback client id."""
    global _default_client_id
    _default_client_id = (client_id or "").strip()
    if _default_client_id:
        _backfill_client_and_user(client_id=_default_client_id)



def _ensure_data_dir() -> None:
    Path(DATA_DIR).mkdir(parents=True, exist_ok=True)


def _create_new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_ACTIVITY_SHEET
    ws.append(ACTIVITY_HEADERS)
    _format_header_row(ws, ACTIVITY_HEADERS)
    _set_column_widths(ws, ACTIVITY_HEADERS)
    if EXCEL_LOCAL_STORAGE_SHEET not in wb.sheetnames:
        local_ws = wb.create_sheet(EXCEL_LOCAL_STORAGE_SHEET)
        local_ws.append(LOCAL_STORAGE_HEADERS)
    
    # Save with proper error handling
    max_attempts = 3
    for attempt in range(max_attempts):
        try:
            # Close file if it exists and is locked
            if EXCEL_PATH.exists():
                try:
                    # Try to open and close to check if locked
                    test_wb = load_workbook(EXCEL_PATH, read_only=True)
                    test_wb.close()
                except Exception:
                    # File is locked, try to delete
                    try:
                        EXCEL_PATH.unlink()
                    except Exception:
                        if attempt < max_attempts - 1:
                            time.sleep(1)
                            continue
            
            wb.save(EXCEL_PATH)
            break
        except PermissionError:
            if attempt < max_attempts - 1:
                time.sleep(1)
            else:
                # Last attempt - save to temp and let user know
                temp_path = EXCEL_PATH.with_suffix('.tmp.xlsx')
                try:
                    wb.save(temp_path)
                except Exception:
                    pass
        except Exception:
            if attempt < max_attempts - 1:
                time.sleep(0.5)
            else:
                pass
    return wb


def _ensure_workbook():
    _ensure_data_dir()
    if EXCEL_PATH.exists():
        try:
            # Try to load with data_only=False to ensure formulas work
            wb = load_workbook(EXCEL_PATH, data_only=False, keep_links=False)
            # Verify it has the required sheets
            if EXCEL_ACTIVITY_SHEET not in wb.sheetnames:
                # Missing required sheet, recreate
                wb.close()
                EXCEL_PATH.unlink(missing_ok=True)
                return _create_new_workbook()
            return wb
        except (BadZipFile, InvalidFileException, KeyError, OSError, IOError) as e:
            # File is corrupted or can't be read, backup and recreate
            try:
                backup_path = EXCEL_PATH.with_suffix('.corrupted.xlsx')
                if EXCEL_PATH.exists():
                    EXCEL_PATH.rename(backup_path)
            except Exception:
                try:
                    EXCEL_PATH.unlink(missing_ok=True)
                except Exception:
                    pass
            return _create_new_workbook()
        except Exception as e:
            # Any other error, try to recreate
            try:
                backup_path = EXCEL_PATH.with_suffix('.error.xlsx')
                if EXCEL_PATH.exists():
                    EXCEL_PATH.rename(backup_path)
            except Exception:
                try:
                    EXCEL_PATH.unlink(missing_ok=True)
                except Exception:
                    pass
            return _create_new_workbook()
    return _create_new_workbook()


def _set_column_widths(ws, headers: list[str]):
    """Set appropriate column widths for better visibility."""
    # Column width mappings (in characters) - significantly increased for clear visibility
    width_map = {
        "client_id": 40,
        "user_id": 40,
        "tool": 50,
        "action": 25,
        "duration": 20,
        "start_time": 30,
        "end_time": 30,
        "task_id": 25,
        "project_id": 25,
        "activity_type": 25,
        "title": 60,
        "metadata_json": 150,
        "screenshots": 120,
        "webcam_photo": 120,
        "webcam_name": 50,
        "mouse_clicks": 18,
        "keys_count": 18,
    }
    for idx, header in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=idx).column_letter
        width = width_map.get(header, 30)  # Default width 30
        ws.column_dimensions[col_letter].width = width


def _format_header_row(ws, headers: list[str]):
    """Format the header row with bold font and background color."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.value = header
    
    # Set row height for header
    ws.row_dimensions[1].height = 25


def _ensure_sheet(wb, sheet_name: str, headers: list[str]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0:
            ws.append(headers)
            _format_header_row(ws, headers)
            _set_column_widths(ws, headers)
            return ws
        header_cells = [cell.value for cell in ws[1]]
        normalized_header = [(cell or "").strip() if isinstance(cell, str) else cell for cell in header_cells]
        
        # Check if we need to add missing columns
        if len(normalized_header) < len(headers) or normalized_header[: len(headers)] != headers:
            # If headers don't match or are missing columns, update them
            if normalized_header[: len(headers)] != headers:
                # Save existing data
                existing_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        existing_data.append(list(row))
                
                # Delete all rows
                ws.delete_rows(1, ws.max_row)
                
                # Add new headers
                ws.append(headers)
                _format_header_row(ws, headers)
                _set_column_widths(ws, headers)
                
                # Restore data with proper column mapping
                for row_data in existing_data:
                    new_row = [""] * len(headers)
                    for idx, val in enumerate(row_data):
                        if idx < len(normalized_header) and idx < len(headers):
                            old_header = normalized_header[idx]
                            if old_header in headers:
                                new_idx = headers.index(old_header)
                                new_row[new_idx] = val
                    ws.append(new_row)
            else:
                # Headers match but might need formatting/width updates
                _format_header_row(ws, headers)
                _set_column_widths(ws, headers)
        else:
            # Ensure column widths and formatting are set even if headers match
            _format_header_row(ws, headers)
            _set_column_widths(ws, headers)
        return ws
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    _format_header_row(ws, headers)
    _set_column_widths(ws, headers)
    return ws


def _excel_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _parse_timestamp(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None
    return None


def _calculate_duration(start_time_str: str, end_time_str: str) -> str:
    """Calculate duration between start_time and end_time in HH:MM:SS format."""
    if not start_time_str or not end_time_str:
        return ""
    try:
        start_dt = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
        end_dt = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")
        delta = end_dt - start_dt
        total_seconds = int(delta.total_seconds())
        if total_seconds < 0:
            return ""
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    except Exception:
        return ""


def _normalize_row(row: Dict[str, Any]) -> Dict[str, Any]:
    normalized: Dict[str, Any] = {key: "" for key in ACTIVITY_HEADERS}
    normalized.update(row or {})
    
    # Set start_time if not present
    if not normalized.get("start_time"):
        normalized["start_time"] = _excel_timestamp()
    
    # Calculate duration if both start_time and end_time are present
    start_time = normalized.get("start_time", "")
    end_time = normalized.get("end_time", "")
    if start_time and end_time and not normalized.get("duration"):
        normalized["duration"] = _calculate_duration(start_time, end_time)
    elif not normalized.get("duration"):
        normalized["duration"] = ""
    
    metadata = normalized.get("metadata_json")
    if isinstance(metadata, (dict, list)):
        normalized["metadata_json"] = json.dumps(metadata, ensure_ascii=False)
    elif metadata is None:
        normalized["metadata_json"] = ""
    screenshot_val = normalized.get("screenshots")
    if isinstance(screenshot_val, (dict, list)):
        normalized["screenshots"] = json.dumps(screenshot_val, ensure_ascii=False)
    elif screenshot_val is None:
        normalized["screenshots"] = ""
    webcam_photo_val = normalized.get("webcam_photo")
    if isinstance(webcam_photo_val, (dict, list)):
        normalized["webcam_photo"] = json.dumps(webcam_photo_val, ensure_ascii=False)
    elif webcam_photo_val is None:
        normalized["webcam_photo"] = ""
    if normalized.get("webcam_name") is None:
        normalized["webcam_name"] = ""
    
    # Ensure mouse_clicks and keys_count are integers
    if normalized.get("mouse_clicks") is None:
        normalized["mouse_clicks"] = 0
    else:
        try:
            normalized["mouse_clicks"] = int(normalized["mouse_clicks"])
        except (ValueError, TypeError):
            normalized["mouse_clicks"] = 0
    
    if normalized.get("keys_count") is None:
        normalized["keys_count"] = 0
    else:
        try:
            normalized["keys_count"] = int(normalized["keys_count"])
        except (ValueError, TypeError):
            normalized["keys_count"] = 0

    if not normalized.get("client_id") and _default_client_id:
        normalized["client_id"] = _default_client_id

    return normalized


def _embed_image_in_cell(
    ws,
    row_num: int,
    col_letter: str,
    artifact: CaptureArtifact | None,
    max_width: int = 200,
    max_height: int = 150,
) -> None:
    """Embed an in-memory image in a specific Excel cell."""
    if not artifact or not artifact.data:
        return
    try:
        pil_image = PILImage.open(BytesIO(artifact.data))
        img = Image(pil_image)
        if img.width <= 0 or img.height <= 0:
            return

        scale_w = max_width / img.width if img.width > max_width else 1.0
        scale_h = max_height / img.height if img.height > max_height else 1.0
        scale = min(scale_w, scale_h)

        new_width = max(50, int(img.width * scale))
        new_height = max(50, int(img.height * scale))

        img.width = new_width
        img.height = new_height
        img.anchor = f"{col_letter}{row_num}"
        ws.add_image(img)

        image_height_points = new_height * 0.75  # Convert pixels to points
        ws.row_dimensions[row_num].height = max(120, image_height_points + 20)
    except Exception as e:
        try:
            from utils.logger import logger

            logger.warning(f"Could not embed image {artifact.filename}: {e}")
        except Exception:
            pass


def _save_workbook_with_retry(wb, max_retries: int = 3, retry_delay: float = 0.5) -> bool:
    """Save workbook with retry logic."""
    # Try to save directly first (faster and more reliable)
    for attempt in range(max_retries):
        try:
            wb.save(EXCEL_PATH)
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
            else:
                # Last attempt - try background save as fallback
                try:
                    _save_executor.submit(_background_save_worker, wb, EXCEL_PATH)
                    return True
                except Exception:
                    return False
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
            else:
                try:
                    from utils.logger import logger
                    logger.error(f"Failed to save workbook after {max_retries} attempts: {e}")
                except Exception:
                    pass
                return False
    return False


def append_activity_event(
    row: Dict[str, Any],
    screenshot_artifact: CaptureArtifact | None = None,
    webcam_artifact: CaptureArtifact | None = None,
) -> None:
    global _last_save_time, _pending_rows, _last_activity_log_time
    normalized = _normalize_row(row)
    
    # Check if we should save immediately (has images) or can batch it
    has_images = bool(
        (screenshot_artifact and screenshot_artifact.has_payload())
        or (webcam_artifact and webcam_artifact.has_payload())
    )
    current_time = time.time()
    # Save more frequently - every 10 seconds instead of 60
    should_save_now = has_images or (current_time - _last_save_time) >= 10.0
    
    # If no images and not time to save, queue for batch write
    if not should_save_now and not has_images:
        with _lock:
            _pending_rows.append((normalized, screenshot_artifact, webcam_artifact))
        return
    
    with _lock:
        try:
            wb = _ensure_workbook()
            ws = _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
            
            # Update previous row's end_time and calculate duration if needed
            if ws.max_row > 1:  # If there are existing rows
                prev_row_num = ws.max_row
                prev_start_time = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("start_time") + 1).value
                prev_end_time = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("end_time") + 1).value
                prev_duration = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("duration") + 1).value
                
                placeholder_end = (
                    prev_start_time
                    and (
                        not prev_end_time
                        or str(prev_end_time) == str(prev_start_time)
                    )
                )
                
                # If previous row doesn't have a real end_time, set it to current start_time
                if prev_start_time and placeholder_end:
                    current_start_time = normalized.get("start_time", _excel_timestamp())
                    ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("end_time") + 1).value = current_start_time
                    
                    # Calculate and set duration for previous row
                    if not prev_duration or prev_duration == "" or (isinstance(prev_duration, (int, float)) and prev_duration < 24):
                        duration = _calculate_duration(str(prev_start_time), current_start_time)
                        duration_cell = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("duration") + 1)
                        duration_cell.value = duration
                        # Set as text format to prevent Excel from converting to number
                        duration_cell.number_format = '@'  # Text format
            
            # Set end_time for current row (will be updated when next activity starts)
            current_start = normalized.get("start_time", _excel_timestamp())
            if not normalized.get("end_time"):
                normalized["end_time"] = ""
            
            ws.append([normalized.get(h, "") for h in ACTIVITY_HEADERS])
            row_num = ws.max_row
            
            # Set duration cell format to text to ensure HH:MM:SS format is preserved
            duration_col_idx = ACTIVITY_HEADERS.index("duration") + 1
            duration_cell = ws.cell(row=row_num, column=duration_col_idx)
            if duration_cell.value:
                duration_cell.number_format = '@'  # Text format
            
            # Embed images in appropriate columns
            if screenshot_artifact and screenshot_artifact.has_payload():
                screenshot_col = get_column_letter(ACTIVITY_HEADERS.index("screenshots") + 1)
                _embed_image_in_cell(
                    ws,
                    row_num,
                    screenshot_col,
                    screenshot_artifact,
                    max_width=300,
                    max_height=200,
                )
                screenshot_cell = ws.cell(row=row_num, column=ACTIVITY_HEADERS.index("screenshots") + 1)
                stored_value = normalized.get("screenshots")
                screenshot_cell.value = stored_value or screenshot_cell.value or screenshot_artifact.filename
            
            if webcam_artifact and webcam_artifact.has_payload():
                webcam_col = get_column_letter(ACTIVITY_HEADERS.index("webcam_photo") + 1)
                _embed_image_in_cell(
                    ws,
                    row_num,
                    webcam_col,
                    webcam_artifact,
                    max_width=300,
                    max_height=200,
                )
                webcam_cell = ws.cell(row=row_num, column=ACTIVITY_HEADERS.index("webcam_photo") + 1)
                stored_webcam_value = normalized.get("webcam_photo")
                webcam_cell.value = stored_webcam_value or webcam_cell.value or webcam_artifact.filename
            
            # Re-apply column widths to ensure they're maintained
            _set_column_widths(ws, ACTIVITY_HEADERS)
            
            # Process pending rows if any
            if _pending_rows:
                for pending_normalized, pending_screenshot, pending_webcam in _pending_rows:
                    # Update previous row's end_time before adding new row
                    if ws.max_row > 1:
                        prev_row_num = ws.max_row
                        prev_start_time = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("start_time") + 1).value
                        prev_end_time = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("end_time") + 1).value
                        prev_duration = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("duration") + 1).value
                        
                        placeholder_end = (
                            prev_start_time
                            and (
                                not prev_end_time
                                or str(prev_end_time) == str(prev_start_time)
                            )
                        )
                        
                        if prev_start_time and placeholder_end:
                            current_start_time = pending_normalized.get("start_time", _excel_timestamp())
                            ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("end_time") + 1).value = current_start_time
                            
                            if not prev_duration or prev_duration == "" or (isinstance(prev_duration, (int, float)) and prev_duration < 24):
                                duration = _calculate_duration(str(prev_start_time), current_start_time)
                                duration_cell = ws.cell(row=prev_row_num, column=ACTIVITY_HEADERS.index("duration") + 1)
                                duration_cell.value = duration
                                duration_cell.number_format = '@'  # Text format
                    
                    # Set end_time for current pending row
                    if not pending_normalized.get("end_time"):
                        pending_normalized["end_time"] = ""
                    
                    ws.append([pending_normalized.get(h, "") for h in ACTIVITY_HEADERS])
                    pending_row_num = ws.max_row
                    # Note: Pending rows don't have images, so skip image embedding
                _pending_rows.clear()
            
            # Always save to ensure data is persisted
            if should_save_now:
                if _save_workbook_with_retry(wb):
                    _last_save_time = current_time
            else:
                # Even if not time to save, ensure pending data is written
                # Save immediately for critical data
                _save_workbook_with_retry(wb)
                _last_save_time = current_time
        except Exception as e:
            # Don't crash on errors, just log
            try:
                from utils.logger import logger
                logger.error(f"Error in append_activity_event: {e}")
            except Exception:
                pass


def append_activity_events(rows: list[Dict[str, Any]]) -> None:
    if not rows:
        return
    with _lock:
        wb = _ensure_workbook()
        ws = _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
        for row in rows:
            normalized = _normalize_row(row)
            ws.append([normalized.get(h, "") for h in ACTIVITY_HEADERS])
        # Re-apply column widths to ensure they're maintained
        _set_column_widths(ws, ACTIVITY_HEADERS)
        wb.save(EXCEL_PATH)


def finalize_last_activity_row(end_time: str | None = None) -> None:
    """
    Ensure the last activity row has an end_time/duration value (used during clock-out).
    """
    with _lock:
        wb = _ensure_workbook()
        ws = _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
        if ws.max_row <= 1:
            return
        row_num = ws.max_row
        start_time_cell = ws.cell(row=row_num, column=ACTIVITY_HEADERS.index("start_time") + 1)
        end_time_cell = ws.cell(row=row_num, column=ACTIVITY_HEADERS.index("end_time") + 1)
        duration_cell = ws.cell(row=row_num, column=ACTIVITY_HEADERS.index("duration") + 1)
        start_time = start_time_cell.value
        if not start_time:
            return
        existing_end = end_time_cell.value
        if end_time:
            end_value = end_time
        elif not existing_end or str(existing_end) == str(start_time):
            end_value = _excel_timestamp()
        else:
            end_value = existing_end
        end_time_cell.value = end_value
        duration = _calculate_duration(str(start_time), str(end_value))
        if duration:
            duration_cell.value = duration
            duration_cell.number_format = '@'
        _save_workbook_with_retry(wb)


def purge_activity_before(cutoff_dt: datetime) -> int:
    """
    Delete Excel activity rows whose start_time is older than cutoff_dt.
    Returns number of deleted rows.
    """
    deleted = 0
    with _lock:
        wb = _ensure_workbook()
        ws = _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
        start_idx = ACTIVITY_HEADERS.index("start_time") + 1
        for row_num in range(ws.max_row, 1, -1):
            start_value = ws.cell(row=row_num, column=start_idx).value
            start_dt = _parse_timestamp(start_value)
            if start_dt and start_dt < cutoff_dt:
                ws.delete_rows(row_num)
                deleted += 1
        if deleted:
            _save_workbook_with_retry(wb)
    return deleted


def make_hyperlink(path_or_url: str, label: str = "Open screenshot (3 files)") -> str:
    if not path_or_url:
        return ""
    quoted = path_or_url.replace('"', '""')
    label_text = label.replace('"', '""')
    return f'=HYPERLINK("{quoted}", "{label_text}")'


def _sheet_rows(ws) -> list[Dict[str, Any]]:
    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        results.append({col: row[idx] for idx, col in enumerate(ACTIVITY_HEADERS)})
    return results


def read_activity_by_date(date_prefix: str) -> list[Dict[str, Any]]:
    """
    Return all activity rows whose start_time begins with the provided YYYY-MM-DD prefix.
    """
    if not date_prefix:
        return []
    with _lock:
        wb = _ensure_workbook()
        ws = _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
        matches = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            start_time = row[ACTIVITY_HEADERS.index("start_time")]
            if isinstance(start_time, str) and start_time.startswith(date_prefix):
                matches.append({col: row[idx] for idx, col in enumerate(ACTIVITY_HEADERS)})
        return matches


def summarize_activity() -> Dict[str, Any]:
    """
    Aggregate total duration per activity_type and counts.
    Duration must be HH:MM:SS string; falls back to 0 when blank.
    """
    def _seconds_from_duration(value: Any) -> int:
        if not value:
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        parts = str(value).split(":")
        try:
            parts = [int(p) for p in parts]
        except ValueError:
            return 0
        while len(parts) < 3:
            parts.insert(0, 0)
        h, m, s = parts
        return h * 3600 + m * 60 + s

    summary: Dict[str, Any] = {"total_rows": 0, "activity_types": {}}
    with _lock:
        wb = _ensure_workbook()
        ws = _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            summary["total_rows"] += 1
            activity_type = row[ACTIVITY_HEADERS.index("activity_type")] or "unknown"
            duration = _seconds_from_duration(row[ACTIVITY_HEADERS.index("duration")])
            bucket = summary["activity_types"].setdefault(activity_type, {"count": 0, "duration_seconds": 0})
            bucket["count"] += 1
            bucket["duration_seconds"] += duration
    return summary


def read_local_storage() -> Dict[str, Any]:
    with _lock:
        wb = _ensure_workbook()
        ws = _ensure_sheet(wb, EXCEL_LOCAL_STORAGE_SHEET, LOCAL_STORAGE_HEADERS)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            key, raw, _ = row[:3]
            if key == "data" and raw:
                try:
                    return json.loads(raw)
                except Exception:
                    return {}
        return {}


def write_local_storage(data: Dict[str, Any]) -> None:
    payload = json.dumps(data or {}, ensure_ascii=False)
    with _lock:
        wb = _ensure_workbook()
        ws = _ensure_sheet(wb, EXCEL_LOCAL_STORAGE_SHEET, LOCAL_STORAGE_HEADERS)
        # Remove previous rows except header
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        ws.append(["data", payload, _excel_timestamp()])
        wb.save(EXCEL_PATH)


def fix_existing_excel_file() -> None:
    """
    Utility function to fix existing Excel file by updating headers, column widths, and formatting.
    This should be called once to migrate old Excel files to the new format.
    """
    with _lock:
        if not EXCEL_PATH.exists():
            return
        try:
            wb = load_workbook(EXCEL_PATH)
            if EXCEL_ACTIVITY_SHEET in wb.sheetnames:
                ws = wb[EXCEL_ACTIVITY_SHEET]
                # Force update headers and formatting
                _ensure_sheet(wb, EXCEL_ACTIVITY_SHEET, ACTIVITY_HEADERS)
                wb.save(EXCEL_PATH)
        except Exception:
            pass
